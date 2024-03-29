import openpyxl
import os
import csv
import pandas as pd

# 确定文件数据所在根目录
src_path = "../data"
# 确定文件保存的新目录
save_src_path = "../cleanedData"
save_csv_path = "../csv_cleanedata"

def cleanRedundantData(new_wholePath):
    # 明确要处理的文件
    print(new_wholePath)
    # 打开文件
    file = pd.read_csv(new_wholePath)
    print("已经打开文件")
    print(file)
    print(file.shape)
    indexID = file[(file['学号/邮箱、电话'] <= 202100000000)].index
    file.drop(indexID, inplace=True)
    print("---------------------")
    print(file)
    info = pd.DataFrame(file)
    csv_data = info.to_csv(new_wholePath)
    print("冗余数据清除完毕")

def xlsxTocsv(save_src_path,save_path,ExperimentalResults_name,save_csv_path,ExperimentalResults_name_csv):
    # 文件所处的原目录完整路径
    allPath = save_src_path + save_path + ExperimentalResults_name
    print(allPath)
    # 打开文件
    workbook = openpyxl.load_workbook(allPath)
    new_wholePath = save_csv_path + save_path + ExperimentalResults_name_csv
    # 确定要更改的sheet目录
    sheet = workbook.active
    # csv_data = []
    #
    # for value in sheet.iter_rows(values_only=True):
    #     csv_data.append(list(value))
    #
    # with open(new_wholePath,'w',encoding='utf-8-sig') as f:
    #     writer = csv.writer(f)
    #     for line in csv_data:
    #         writer.writerow(line)

    with open(new_wholePath, "w", encoding="utf-8-sig",newline='') as f:
        write = csv.writer(f)
        data = []
        for i in range(3, sheet.max_row + 1):
            row_stack = []
            for j in range(1, sheet.max_column + 1):
                row_stack.append(sheet.cell(row=i, column=j).value)
            data.append(row_stack)
        write.writerows(data)

    # with open(new_wholePath,'r') as f:
    #     for line in f:
    #         if line.strip():
    #             print(line)

    # pd_reader = pd.read_csv(new_wholePath)
    # print(pd_reader)
    cleanRedundantData(new_wholePath)


def openFile(src_path,ExperimentalResults_path,ExperimentalResults_name,save_src_path,save_path,ExperimentalResults_name_csv):
    # 获取查重文件的完整路径
    wholePath = src_path + ExperimentalResults_path + ExperimentalResults_name;
    print("完整路径为："+wholePath)
    # 使用openpyxl加载文件
    file = openpyxl.load_workbook(wholePath)
    print("已加载文件")
    # 根据sheet名称获取要操作的指定sheet页
    sheet = file['Sheet0']
    print("指定sheet页")
    # 读取表内所有数据
    data_of_row = list(sheet.rows)
    # for row in sheet.rows:
    #     for cell in row:
    #         print(cell.value)
    file_act = file.active
    # 获取该文件中的最大行数
    file_rows_max_nums = file_act.max_row
    # 起始行，由于存在3行表头，故为3
    row_n = 4
    # 目标列，为第一列，主要为了取出多余数据
    col_n = 1
    i = 3
    while i < file_rows_max_nums:
        if file_act.cell(row=row_n,column=col_n).value != '21级12班':
            file_act.delete_rows(row_n)
            print("已删除第"+str(row_n)+"行")
        else:
            row_n += 1
        i += 1
    save_name = save_src_path + save_path + ExperimentalResults_name
    print(save_name)
    file.save(save_name)
    print("已成功保存"+ExperimentalResults_name)
    # 将xlsx文件转换为CSV文件，以便后续导入数据库文件
    xlsxTocsv(save_src_path,save_path,ExperimentalResults_name,save_csv_path,ExperimentalResults_name_csv)


# 进行实验查重文件的清洗过程
# 确定实验查重文件所在的目录
ExperimentalResults_path = "/实验成绩"
# 要保存的路径
save_path = "/ExperimentalResults"

# 确定文件名称，并访问打开文件从而进行数据清洗
for i in range(1,7):
    # 要查重文件的名称
    ExperimentalResults_name = "/21级12班《数据结构与算法》第"+str(i)+"次实验课_成绩单"+".xlsx"
    ExperimentalResults_name_csv = "/21级12班《数据结构与算法》第"+str(i)+"次实验课_成绩单"+".csv"
    # 打印要查重文件的名称
    print(ExperimentalResults_name)
    # 从系统中根据文件目录打开文件
    openFile(src_path,ExperimentalResults_path,ExperimentalResults_name,save_src_path,save_path,ExperimentalResults_name_csv)
