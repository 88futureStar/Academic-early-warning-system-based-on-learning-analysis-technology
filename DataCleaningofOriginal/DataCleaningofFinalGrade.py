import openpyxl
import pandas as pd
import csv


# 确定文件数据所在根目录
src_path = "../data"
# 确定文件保存的新目录
save_src_path = "../cleanedData"
# 转化为CSV文件保存的新目录
save_csv_path = "../csv_cleanedata"

def ToHundred(save_csv_path,save_path,FinalGrade_name_csv):
    path = save_csv_path + save_path + FinalGrade_name_csv
    print(path)
    df = pd.read_csv(path)
    # 显示在运行控制台上
    print(df)
    # 将其存储的字段转为字符串类型，以便后续进行相关处理
    # CSV文件中的数据以字符串格式存储，并以，进行分隔，因此不用转换数据类型
    # 选择其中相对应的期末成绩的总分一列，根据总分将其折合为百分制
    # score_hundred_orginaldata用于存储原先的总分数据
    score_hundred_orginaldata = []

    # 需要运行者提供总分
    print("请输入本次考试的总分为：")
    # 从控制台获取用户输入的总分
    total_score = input()
    print(total_score)

    for index, row in df.iterrows():
        # score_hundred_orginaldata.append()
        # print(row[3])
        # 将所有的总分存入score_hundred_orginaldata，以待后续进行处理
        # 其中有些同学未参加考试，显示缺考，这一部分数据将改为0分处理
        if row[3] == "缺考":
            score_hundred_orginaldata.append(0)
        # 其余则将字符串转为整形并存储在score_hundred_orginaldata当中
        else:
            score_hundred_orginaldata.append(int(row[3]))
    print(score_hundred_orginaldata)

    # score_hundred用于存储处理后的数据
    score_hundred = []

    # score_rank用于存储折合后的成绩的排名
    score_rank = []

    # score_class用于存储排名的类型
    score_class = ['A', 'B', 'C', 'D', 'F']

    for i in range(len(score_hundred_orginaldata)):
        print(score_hundred_orginaldata[i])
        # 将总分折合为百分制
        score_handled = score_hundred_orginaldata[i] * 100 / int(total_score)
        print(score_handled)
        score_hundred.append(score_handled)
        if score_handled > 90:
            print(score_class[0])
            score_rank.append(score_class[0])
        elif score_handled > 80:
            print(score_class[1])
            score_rank.append(score_class[1])
        elif score_handled > 70:
            print(score_class[2])
            score_rank.append(score_class[2])
        elif score_handled > 60:
            print(score_class[3])
            score_rank.append(score_class[3])
        else:
            print(score_class[4])
            score_rank.append(score_class[4])
        print("------------------------------------")

    print(score_rank)
    message = open(path, 'r', encoding='utf-8')
    # 1.2创建读文件方法
    message_1 = csv.reader(message)
    # 将数据类型变成有序列表
    data = list(message_1)
    message.close()  # 关闭文件
    f = open(path, 'w', encoding='utf-8', newline='')
    mywriter = csv.writer(f)
    data1 = [i for i in data[0]]
    data1.append('成绩类别')
    mywriter.writerow((data1))
    li = 0
    for i in data[1:]:
        # print(i)
        # avg = round((float(i[3]) + float(i[4]) + float(i[5]) + float(i[6])) / 4, 2)
        # result = f'{i[0]}的平均分为：{avg}'
        # print(result)
        # 将平均分重新写入文件
        print(score_rank[li])
        i.append(score_rank[li])
        mywriter.writerow(i)
        li = li + 1
    f.close()
    print("数据修改完毕")


def xlsxTocsv(save_src_path,save_path,FinalGrade_name,save_csv_path,FinalGrade_name_csv):
    # 文件所处的原目录完整路径
    allPath = save_src_path + save_path + FinalGrade_name
    print(allPath)
    # 打开文件
    workbook = openpyxl.load_workbook(allPath)
    new_wholePath = save_csv_path + save_path + FinalGrade_name_csv
    # 确定要更改的sheet目录
    sheet = workbook.active
    # csv_data = []
    #
    # for value in sheet.iter_rows(values_only=True):
    #     csv_data.append(list(value))
    #
    # with open(new_wholePath,'w',encoding='utf-8-sig') as f:
    #     writer = csv.writer(f)
    #     for line in csv_data:
    #         writer.writerow(line)

    with open(new_wholePath, "w", encoding="utf-8-sig",newline='') as f:
        write = csv.writer(f)
        data = []
        for i in range(3, sheet.max_row + 1):
            row_stack = []
            for j in range(1, sheet.max_column + 1):
                row_stack.append(sheet.cell(row=i, column=j).value)
            data.append(row_stack)
        write.writerows(data)

    # with open(new_wholePath,'r') as f:
    #     for line in f:
    #         if line.strip():
    #             print(line)

    # pd_reader = pd.read_csv(new_wholePath)
    # print(pd_reader)

def openFile(src_path,FinalGrade_path,FinalGrade_name,save_src_path,save_path,FinalGrade_name_csv):
    # 获取查重文件的完整路径
    wholePath = src_path + FinalGrade_path + FinalGrade_name;
    print("完整路径为："+wholePath)
    # 使用openpyxl加载文件
    file = openpyxl.load_workbook(wholePath)
    print("已加载文件")
    # 根据sheet名称获取要操作的指定sheet页
    sheet = file['Sheet0']
    print("指定sheet页")
    # 读取表内所有数据
    data_of_row = list(sheet.rows)
    # for row in sheet.rows:
    #     for cell in row:
    #         print(cell.value)
    file_act = file.active
    # 获取该文件中的最大行数
    file_rows_max_nums = file_act.max_row
    # 起始行，由于存在3行表头，故为3
    # row_n = 5
    # # 目标列，为第一列，主要为了取出多余数据
    # col_n = 1
    # i = 3
    # while i < file_rows_max_nums:
    #     if file_act.cell(row=row_n,column=col_n).value != '21级12班':
    #         file_act.delete_rows(row_n)
    #         print("已删除第"+str(row_n)+"行")
    #     else:
    #         row_n += 1
    #     i += 1
    save_name = save_src_path + save_path + FinalGrade_name
    print(save_name)
    file.save(save_name)
    print("已成功保存"+FinalGrade_name)
    # 将xlsx文件转换为CSV文件，以便后续导入数据库文件
    xlsxTocsv(save_src_path,save_path,FinalGrade_name,save_csv_path,FinalGrade_name_csv)
    # 对转换为CSV文件的数据的总分进行折合
    ToHundred(save_csv_path,save_path,FinalGrade_name_csv)

# 进行实验查重文件的清洗过程
# 确定实验查重文件所在的目录
FinalGrade_path = "/期末成绩"
# 要保存的路径
save_path = "/FinalGrade"

# 确定文件名称，并访问打开文件从而进行数据清洗
for i in range(1,2):
    # 要查重文件的名称
    FinalGrade_name = "/2021级软件工程专业《数据结构与算法》期末考试-成绩单.xlsx"
    FinalGrade_name_csv = "/2021级软件工程专业《数据结构与算法》期末考试-成绩单.csv"
    # 打印要查重文件的名称
    print(FinalGrade_name)
    # 从系统中根据文件目录打开文件
    openFile(src_path,FinalGrade_path,FinalGrade_name,save_src_path,save_path,FinalGrade_name_csv)
